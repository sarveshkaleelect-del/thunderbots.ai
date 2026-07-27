"""
ThunderBots Smart Shop Assistant — Sync Service (NEW)

The database (shop_assistant_products) is always the source of truth.
Excel and Google Sheets are editable VIEWS onto it, never a parallel store:

- export_to_xlsx() / import_from_xlsx()   Always synchronous, in-memory,
    zero external dependency — works with no configuration at all.
- push_to_google_sheet() / pull_from_google_sheet()   Require a shop owner
    to have connected a Google service-account credential (ShopSyncConfig).
    Fully inert (no network calls, nothing scheduled) until that happens.

Conflict handling: both directions are "last write wins" at the row level,
keyed by product name (case-insensitive). Pulling from a sheet that has a
row the DB doesn't know about CREATES a new product; a DB product missing
from the sheet is left untouched (never silently deleted) — deletion must
be an explicit action in the admin UI, never an implicit side effect of a
sync.
"""
import io
import json
import logging
from datetime import datetime, timezone

from cryptography.fernet import InvalidToken
from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.shop_assistant import ShopProduct, ShopSyncConfig
from app.services.ai_engine import encrypt_key, decrypt_key  # reused as-is, not modified

logger = logging.getLogger(__name__)

_HEADER = ["Product Name", "SKU", "Category", "Available Quantity"]


class SyncConfigError(Exception):
    pass


# ─────────────────────────────────────────────────────────────────────────────
# Excel
# ─────────────────────────────────────────────────────────────────────────────

def export_to_xlsx(products: list[ShopProduct]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventory"
    ws.append(_HEADER)
    for p in products:
        ws.append([p.name, p.sku or "", p.category or "", p.quantity_available])
    for col_idx, width in enumerate((36, 16, 16, 18), start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_xlsx(file_bytes: bytes) -> list[dict]:
    """Parses an uploaded .xlsx into row dicts. Raises ValueError on a file
    that doesn't look like a valid inventory export (missing header)."""
    try:
        wb = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as e:
        raise ValueError(f"Could not read Excel file: {e}") from e

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []

    header = [str(c).strip() if c else "" for c in rows[0]]
    expected = {"product name", "available quantity"}
    if not expected.issubset({h.lower() for h in header}):
        raise ValueError(
            "Expected columns 'Product Name' and 'Available Quantity' "
            "(optionally 'SKU' and 'Category') were not found in row 1."
        )

    idx = {h.lower(): i for i, h in enumerate(header)}
    parsed = []
    for row in rows[1:]:
        if row is None or all(c is None for c in row):
            continue
        name = row[idx["product name"]] if idx.get("product name") is not None else None
        if not name or not str(name).strip():
            continue
        qty_raw = row[idx["available quantity"]] if "available quantity" in idx else 0
        try:
            qty = int(qty_raw) if qty_raw is not None else 0
        except (TypeError, ValueError):
            qty = 0
        parsed.append({
            "name": str(name).strip(),
            "sku": (str(row[idx["sku"]]).strip() if "sku" in idx and row[idx["sku"]] else None),
            "category": (str(row[idx["category"]]).strip() if "category" in idx and row[idx["category"]] else None),
            "quantity_available": max(qty, 0),
        })
    return parsed


async def apply_rows_to_db(db: AsyncSession, shop_id: str, rows: list[dict], *, event_type: str = "import") -> dict:
    """Upserts parsed rows (from Excel or Sheets) into shop_assistant_products,
    keyed by case-insensitive product name. Never deletes a product that's
    simply absent from the incoming rows.

    Every quantity change is written to the movement ledger (event_type
    "import" or "sync_pull") so AI Inventory Intelligence sees it, and the
    returned summary includes `increased_product_ids` so the caller can run
    process_waitlist_for_product() for each one — a bulk import/sync
    restocking a product should notify waiting customers exactly like a
    manual restock does."""
    # Local import to avoid a circular import at module load time (shop_assistant_service
    # does not import shop_sync_service, so this is safe).
    from app.services.shop_assistant_service import record_movement

    result = await db.execute(select(ShopProduct).where(ShopProduct.shop_id == shop_id))
    existing = {p.name.strip().lower(): p for p in result.scalars().all()}

    created, updated = 0, 0
    increased_product_ids: list[str] = []
    for row in rows:
        key = row["name"].strip().lower()
        if key in existing:
            product = existing[key]
            before = product.quantity_available
            new_qty = row["quantity_available"]
            if new_qty != before:
                product.quantity_available = new_qty
                await record_movement(
                    db, shop_id=shop_id, product_id=product.id, event_type=event_type,
                    quantity_delta=new_qty - before, units=abs(new_qty - before),
                    quantity_before=before, quantity_after=new_qty,
                )
                if new_qty > before:
                    increased_product_ids.append(product.id)
            if row.get("sku"):
                product.sku = row["sku"]
            if row.get("category"):
                product.category = row["category"]
            updated += 1
        else:
            product = ShopProduct(
                shop_id=shop_id,
                name=row["name"],
                sku=row.get("sku"),
                category=row.get("category"),
                quantity_available=row["quantity_available"],
            )
            db.add(product)
            await db.flush()
            await record_movement(
                db, shop_id=shop_id, product_id=product.id, event_type=event_type,
                quantity_delta=0, units=0, quantity_before=0, quantity_after=product.quantity_available,
            )
            created += 1

    await db.flush()
    return {
        "created": created, "updated": updated, "total_rows": len(rows),
        "increased_product_ids": increased_product_ids,
    }


# ─────────────────────────────────────────────────────────────────────────────
# Google Sheets
# ─────────────────────────────────────────────────────────────────────────────

def encrypt_credentials(service_account_json: str) -> str:
    # Validate it's actually parseable JSON before we ever store it.
    try:
        json.loads(service_account_json)
    except json.JSONDecodeError as e:
        raise ValueError(f"Not valid JSON: {e}") from e
    return encrypt_key(service_account_json)


def _decrypt_credentials(encrypted: str) -> dict:
    try:
        raw = decrypt_key(encrypted)
    except InvalidToken as e:
        raise SyncConfigError("Stored Google credentials could not be decrypted") from e
    return json.loads(raw)


def _open_worksheet(config: ShopSyncConfig):
    """Lazily imports gspread/google-auth (optional dependency path — only
    ever hit once a shop owner has actually connected Google Sheets) and
    opens the configured worksheet."""
    try:
        import gspread
        from google.oauth2.service_account import Credentials
    except ImportError as e:
        raise SyncConfigError(
            "Google Sheets sync requires the 'gspread' and 'google-auth' "
            "packages to be installed on the server."
        ) from e

    creds_dict = _decrypt_credentials(config.encrypted_credentials)
    scopes = ["https://www.googleapis.com/auth/spreadsheets"]
    credentials = Credentials.from_service_account_info(creds_dict, scopes=scopes)
    client = gspread.authorize(credentials)
    sheet = client.open_by_key(config.spreadsheet_id)
    try:
        return sheet.worksheet(config.worksheet_name)
    except gspread.WorksheetNotFound:
        return sheet.add_worksheet(title=config.worksheet_name, rows=1000, cols=len(_HEADER))


def push_to_google_sheet(config: ShopSyncConfig, products: list[ShopProduct]) -> None:
    """DB -> Sheet. Overwrites the sheet's data rows with the current live
    inventory. Requires the shop owner to have connected a service account
    with edit access to `config.spreadsheet_id` (they must also have shared
    the sheet with the service account's email — same as any Google Sheets
    API integration)."""
    ws = _open_worksheet(config)
    values = [_HEADER] + [
        [p.name, p.sku or "", p.category or "", p.quantity_available] for p in products
    ]
    ws.clear()
    ws.update(values=values, range_name="A1")


def pull_from_google_sheet(config: ShopSyncConfig) -> list[dict]:
    """Sheet -> parsed rows, reusing the exact same row-shape/validation as
    the Excel import path so both sync sources upsert identically via
    apply_rows_to_db()."""
    ws = _open_worksheet(config)
    records = ws.get_all_values()
    if not records:
        return []
    header = [c.strip().lower() for c in records[0]]
    if "product name" not in header or "available quantity" not in header:
        raise SyncConfigError(
            "The connected sheet's header row must contain 'Product Name' "
            "and 'Available Quantity' columns."
        )
    idx = {h: i for i, h in enumerate(header)}
    parsed = []
    for row in records[1:]:
        if not row or not any(c.strip() for c in row):
            continue
        name = row[idx["product name"]].strip() if idx["product name"] < len(row) else ""
        if not name:
            continue
        try:
            qty = int(row[idx["available quantity"]]) if idx["available quantity"] < len(row) else 0
        except ValueError:
            qty = 0
        parsed.append({
            "name": name,
            "sku": row[idx["sku"]].strip() if "sku" in idx and idx["sku"] < len(row) and row[idx["sku"]] else None,
            "category": row[idx["category"]].strip() if "category" in idx and idx["category"] < len(row) and row[idx["category"]] else None,
            "quantity_available": max(qty, 0),
        })
    return parsed


def mark_sync_result(config: ShopSyncConfig, error: str | None) -> None:
    config.last_synced_at = datetime.now(timezone.utc)
    config.last_sync_error = error
