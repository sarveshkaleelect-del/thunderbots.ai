"""
ThunderBots AI Business Advisor — Report Export (NEW)

Generates the same daily/weekly/monthly report payload produced by
business_advisor_service.get_report() as a downloadable PDF (reportlab) or
Excel (openpyxl) file. Purely a rendering layer — no new calculations.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer


def build_xlsx(report: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")

    ws.append([f"AI Business Advisor — {report['period'].capitalize()} Report"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([f"Shop: {report['shop_name']}"])
    ws.append([f"Period: {report['start_date']} to {report['end_date']}"])
    ws.append([])
    ws.append(["Total Revenue", "Total Profit", "Total Orders"])
    for cell in ws[5]:
        cell.font = header_font
        cell.fill = header_fill
    ws.append([report["total_revenue"], report["total_profit"], report["total_orders"]])
    ws.append([])

    headers = ["Date", "Revenue", "Profit", "Orders", "New Customers", "Returning Customers"]
    ws.append(headers)
    header_row = ws.max_row
    for cell in ws[header_row]:
        cell.font = header_font
        cell.fill = header_fill

    for row in report["daily_breakdown"]:
        ws.append([
            row["date"], row["revenue"], row["profit"], row["orders"],
            row.get("new_customers", 0), row.get("returning_customers", 0),
        ])

    for col in ws.columns:
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = max(12, min(28, max_len + 2))

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_pdf(report: dict) -> bytes:
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=1.5 * cm, bottomMargin=1.5 * cm)
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("TBTitle", parent=styles["Title"], textColor=colors.HexColor("#312e81"))
    subtitle_style = ParagraphStyle("TBSubtitle", parent=styles["Normal"], textColor=colors.HexColor("#4b5563"))

    elements = [
        Paragraph(f"AI Business Advisor — {report['period'].capitalize()} Report", title_style),
        Spacer(1, 4),
        Paragraph(f"Shop: {report['shop_name']}", subtitle_style),
        Paragraph(f"Period: {report['start_date']} to {report['end_date']}", subtitle_style),
        Spacer(1, 14),
    ]

    summary_data = [
        ["Total Revenue", "Total Profit", "Total Orders"],
        [f"{report['total_revenue']}", f"{report['total_profit']}", f"{report['total_orders']}"],
    ]
    summary_table = Table(summary_data, colWidths=[5.5 * cm, 5.5 * cm, 5.5 * cm])
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#6366f1")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 8),
    ]))
    elements.append(summary_table)
    elements.append(Spacer(1, 18))

    elements.append(Paragraph("Daily Breakdown", styles["Heading2"]))
    elements.append(Spacer(1, 6))

    table_data = [["Date", "Revenue", "Profit", "Orders", "New", "Returning"]]
    for row in report["daily_breakdown"]:
        table_data.append([
            row["date"], str(row["revenue"]), str(row["profit"]), str(row["orders"]),
            str(row.get("new_customers", 0)), str(row.get("returning_customers", 0)),
        ])

    daily_table = Table(table_data, repeatRows=1)
    daily_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#312e81")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#e5e7eb")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f5f5ff")]),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
    ]))
    elements.append(daily_table)

    doc.build(elements)
    return buf.getvalue()
